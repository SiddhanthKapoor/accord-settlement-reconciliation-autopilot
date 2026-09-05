"""
One way in for tabular uploads, whatever the file actually is.

Finance teams do not send one clean CSV. They send a CSV, a bank's XLSX
export with the account holder's name in row 1 and the real header in
row 4, and a file called `settlements.xlsx` that is a CSV somebody
renamed. All three have to land in the same place, because everything
downstream — schema detection, classification, mapping — is written
against `(columns, rows)` and must not learn a second set of rules per
file format.

Two decisions here are load-bearing:

*Format comes from the bytes, not the extension.* An XLSX is a zip and
starts `PK\\x03\\x04`; anything else that claims to be one is read as
text. Trusting the extension means a renamed file is either rejected or,
worse, parsed as a single unusable column.

*Cell coercion matches the CSV path exactly.* Every value becomes a
string in the same shape a CSV would have carried, and floats keep their
decimal point. That last part is not cosmetic: `detect_amount_scale`
reads a column of whole numbers as paise, so rendering 12500.0 as
"12500" would silently divide a bank statement by a hundred.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from decimal import Decimal

from app.ingest.schema import parse_amount, parse_csv, parse_date

# Well above a month of transactions for any business that would upload
# a file by hand, and low enough that a mis-sent 2GB export cannot take
# the process down. A truncated read is reported, never silent.
MAX_ROWS = 100_000

# How far into a sheet to look for the real header before giving up and
# treating row 1 as the header.
HEADER_SCAN_ROWS = 30

XLSX_MAGIC = b"PK\x03\x04"
XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


class UnreadableFile(ValueError):
    """The file cannot be turned into rows, with a reason a user can act on."""


@dataclass
class ReadResult:
    columns: list[str]
    rows: list[dict]
    fmt: str                       # "csv" | "xlsx"
    csv_text: str                  # canonical CSV, what gets stored and re-parsed
    truncated: bool = False
    header_row: int | None = None  # 1-based row number of the header within the sheet
    sheet_name: str | None = None
    notes: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------

def detect_format(filename: str, raw: bytes) -> tuple[str, list[str]]:
    """`("csv" | "xlsx", notes)` from the bytes, disagreeing with the
    extension out loud when it has to."""
    name = (filename or "").lower()
    claims_xlsx = name.endswith((".xlsx", ".xlsm"))
    notes: list[str] = []

    if raw.startswith(XLS_MAGIC):
        raise UnreadableFile(
            "this is a legacy .xls (Excel 97-2003) workbook, which cannot be read here — "
            "re-save it as .xlsx or CSV"
        )
    if raw.startswith(XLSX_MAGIC):
        if not claims_xlsx:
            notes.append(f"'{filename}' is not named .xlsx but the bytes are an XLSX workbook — read as XLSX")
        return "xlsx", notes
    if claims_xlsx:
        notes.append(f"'{filename}' is named .xlsx but the bytes are not a workbook — read as text/CSV")
    return "csv", notes


def decode_text(raw: bytes) -> str:
    try:
        return raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        return raw.decode("latin-1")


# ---------------------------------------------------------------------------
# Cell coercion
# ---------------------------------------------------------------------------

def cell_to_text(value) -> str:
    """An XLSX cell as the string a CSV would have held.

    `datetime` is checked before `date` because it is a subclass of it,
    and floats go through `repr` rather than a fixed format so that
    12500.0 stays "12500.0" (major units) instead of becoming "12500",
    which downstream scale detection would read as paise.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        if (value.hour, value.minute, value.second, value.microsecond) == (0, 0, 0, 0):
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, timedelta):
        return str(value)
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return repr(value)
    return str(value).strip()


def _non_empty(cells) -> list:
    return [c for c in cells if c is not None and str(c).strip() != ""]


def _parses_as_value(text: str) -> bool:
    return parse_amount(text) is not None or parse_date(text) is not None


def _looks_like_data_row(cells, header_width: int) -> bool:
    """A data row under a header: populated to roughly the header's width,
    and carrying at least one number or date, or else matching the header
    width exactly (an all-text data row is unusual but legitimate)."""
    populated = _non_empty(cells)
    if len(populated) < max(1, header_width // 2):
        return False
    if any(_parses_as_value(cell_to_text(c)) for c in populated):
        return True
    return len(populated) == header_width


def find_header_row(rows: list[tuple], scan: int = HEADER_SCAN_ROWS) -> int:
    """Index of the row that is actually the header.

    Bank and ERP exports routinely open with a title, an account number
    and a blank line. The header is the first row that is mostly
    non-empty text and is followed by something that reads as data;
    anything less specific matches the title line, and then every column
    name in the file is wrong.
    """
    for i in range(min(len(rows), scan)):
        populated = _non_empty(rows[i])
        if len(populated) < 2:
            continue                                   # a title is one wide cell
        stringy = [c for c in populated if isinstance(c, str)]
        if len(stringy) / len(populated) < 0.6:
            continue                                   # mostly numbers: this is data
        # The first populated row after the candidate, allowing for a
        # blank spacer line between header and data.
        following = next((rows[j] for j in range(i + 1, min(i + 4, len(rows))) if _non_empty(rows[j])), None)
        if following is None:
            continue
        if _looks_like_data_row(following, len(populated)):
            return i
    return 0


def _header_names(cells) -> list[str]:
    """Header cells as unique, non-empty column names.

    Blank and duplicated headers are real — a bank export with two
    "Amount" columns is not rare — and collapsing them would drop a
    column's data silently, so they are made unique instead.
    """
    names: list[str] = []
    seen: dict[str, int] = {}
    for i, cell in enumerate(cells):
        name = cell_to_text(cell).strip() or f"column_{i + 1}"
        count = seen.get(name.lower(), 0)
        seen[name.lower()] = count + 1
        names.append(name if count == 0 else f"{name}_{count + 1}")
    return names


def _to_csv_text(columns: list[str], rows: list[dict]) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(columns)
    for row in rows:
        writer.writerow([row.get(c, "") for c in columns])
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Readers
# ---------------------------------------------------------------------------

def _read_xlsx(raw: bytes, *, max_rows: int, sheet: str | None) -> ReadResult:
    try:
        import openpyxl
    except ImportError as exc:                                  # pragma: no cover
        raise UnreadableFile("XLSX support needs openpyxl installed") from exc

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:                                    # noqa: BLE001
        raise UnreadableFile(f"could not open this workbook ({type(exc).__name__})") from exc

    try:
        if sheet is not None:
            if sheet not in workbook.sheetnames:
                raise UnreadableFile(
                    f"sheet '{sheet}' is not in this workbook (found: {', '.join(workbook.sheetnames)})"
                )
            worksheet = workbook[sheet]
        else:
            if not workbook.worksheets:
                raise UnreadableFile("this workbook has no sheets")
            worksheet = workbook.worksheets[0]

        notes: list[str] = []
        if sheet is None and len(workbook.sheetnames) > 1:
            notes.append(
                f"workbook has {len(workbook.sheetnames)} sheets; read '{worksheet.title}' "
                f"(others ignored: {', '.join(n for n in workbook.sheetnames if n != worksheet.title)})"
            )

        # Buffered with room for a header offset so the cap still yields
        # max_rows of *data* when the sheet opens with a title block.
        buffer_limit = max_rows + HEADER_SCAN_ROWS + 1
        raw_rows: list[tuple] = []
        overflow = False
        for row in worksheet.iter_rows(values_only=True):
            if len(raw_rows) >= buffer_limit:
                overflow = True
                break
            raw_rows.append(row)
    finally:
        workbook.close()

    if not raw_rows:
        return ReadResult(columns=[], rows=[], fmt="xlsx", csv_text="", sheet_name=worksheet.title, notes=notes)

    header_index = find_header_row(raw_rows)
    if header_index > 0:
        notes.append(f"header found on row {header_index + 1}; rows above it were a title or blank block")

    columns = _header_names(raw_rows[header_index])
    body = raw_rows[header_index + 1:]
    truncated = overflow or len(body) > max_rows
    body = body[:max_rows]

    rows: list[dict] = []
    for cells in body:
        if not _non_empty(cells):
            continue                                            # blank spacer line
        padded = list(cells) + [None] * (len(columns) - len(cells))
        rows.append({name: cell_to_text(padded[i]) for i, name in enumerate(columns)})

    # Trailing columns that are empty everywhere are Excel's padding, not
    # data. Dropped only when both the header cell and every value are
    # empty, so a genuinely unnamed but populated column survives.
    while columns and columns[-1].startswith("column_") and all(not r.get(columns[-1]) for r in rows):
        dead = columns.pop()
        for row in rows:
            row.pop(dead, None)

    if truncated:
        notes.append(f"read the first {len(rows):,} data rows; the sheet has more")

    return ReadResult(
        columns=columns, rows=rows, fmt="xlsx", csv_text=_to_csv_text(columns, rows),
        truncated=truncated, header_row=header_index + 1, sheet_name=worksheet.title, notes=notes,
    )


def _read_csv(raw: bytes, *, max_rows: int) -> ReadResult:
    text = decode_text(raw)
    # One over the cap so truncation is observed rather than assumed.
    columns, rows = parse_csv(text, limit=max_rows + 1)
    truncated = len(rows) > max_rows
    notes: list[str] = []
    if truncated:
        rows = rows[:max_rows]
        notes.append(f"read the first {len(rows):,} rows; the file has more")
    # An untruncated CSV is stored byte-for-byte as uploaded; only a
    # truncated one is re-serialised, so what is re-parsed at execution
    # is exactly what was read here.
    csv_text = _to_csv_text(columns, rows) if truncated else text
    return ReadResult(columns=columns, rows=rows, fmt="csv", csv_text=csv_text,
                      truncated=truncated, notes=notes)


def read_table(filename: str, raw: bytes, *, max_rows: int = MAX_ROWS, sheet: str | None = None) -> ReadResult:
    """Any supported upload as `(columns, rows)` plus how it was read.

    Raises `UnreadableFile` with a message meant for the person who
    uploaded it. Returning an empty result instead would push the
    decision downstream, where the only available answer is a generic
    "no columns found".
    """
    fmt, notes = detect_format(filename or "", raw)
    result = _read_xlsx(raw, max_rows=max_rows, sheet=sheet) if fmt == "xlsx" else _read_csv(raw, max_rows=max_rows)
    result.notes = notes + result.notes
    return result
