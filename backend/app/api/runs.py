"""
Reconciliation runs over uploaded data.

The flow a finance operator actually performs: create a run, upload the
files they have, confirm what each column means, execute, inspect. The
dataset-driven `/batch/run` path stays exactly as it was — it is how the
frozen evaluations reproduce — and both write into the same records,
review queue and audit tables, so an uploaded run is inspected with the
same screens and the same evidence as an evaluation run.

Column mapping is confirmed rather than assumed. Detection reports a
confidence per column and refuses to proceed while a required field is
unresolved. Silently mis-reading an amount column is the one failure this
product cannot afford.
"""

from __future__ import annotations

import hashlib
import json
import threading
import uuid
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from pydantic import BaseModel

from app.domain.models import PolicyConfig, ReconciliationRecord
from app.engine.batch import process_batch
from app.engine.matching import ReferenceIndex
from app.engine.semantic import get_semantic_verifier
from app.ingest import flow
from app.ingest.classify import classify_source
from app.ingest.mapper import combine, combine_provenance, map_rows
from app.ingest.reader import UnreadableFile, read_table
from app.ingest.schema import CANONICAL_FIELDS, SourceType, detect_schema, parse_csv
from app.ledger import audit, store

router = APIRouter()

# A single upload is held in memory to parse it, so it is bounded. Well
# above a realistic month of transactions, well below anything that
# threatens the process.
MAX_UPLOAD_BYTES = 32 * 1024 * 1024

# One workspace's sources all live in memory during a run, so the
# per-file cap is not enough on its own — fifty files each just under the
# per-file limit is a different problem from one oversized file.
MAX_WORKSPACE_BYTES = 128 * 1024 * 1024
MAX_SOURCES_PER_RUN = 50

# Identifier values kept per source so the plan can propose relationships
# on observed overlap rather than on filenames. Sampled, and labelled as
# sampled everywhere it is used.
IDENTIFIER_SAMPLE_SIZE = 200
IDENTIFIER_SAMPLE_ROWS = 2_000

PREVIEW_ROWS = 200

# `source_type` may be omitted entirely or sent as this sentinel; either
# means "work it out from the file".
AUTO = "AUTO"

# The sample workspace: a folder of month-end exports kept in the repo so
# the product can be tried without finding files first. Read from disk at
# request time — never a hardcoded list, so the folder is the source of
# truth and can grow or shrink without this module knowing.
DEMO_WORKSPACE_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "demo_workspace"

# Sidecar files that sit alongside the exports and are not exports. A
# leading `_` or `.` is the convention for them here; the extensions are
# the metadata formats that would otherwise be read as a one-column CSV.
NON_TABULAR_SUFFIXES = frozenset({".json", ".md", ".py", ".yaml", ".yml", ".toml", ".lock"})


class CreateRunRequest(BaseModel):
    label: str | None = None


class MappingRequest(BaseModel):
    mapping: dict[str, str]
    source_type: str | None = None
    amount_scale: str | None = None


class ExecuteRequest(BaseModel):
    label: str | None = None


class PlanSourceUpdate(BaseModel):
    """A user's answer to "what is this file, and which side is it on?"."""

    source_id: str
    source_type: str | None = None
    role: str | None = None
    confirmed: bool = True
    note: str | None = None


class PlanRelationshipUpdate(BaseModel):
    from_source_id: str
    to_source_id: str
    confirmed: bool = True
    note: str | None = None


class PlanRequest(BaseModel):
    sources: list[PlanSourceUpdate] = []
    relationships: list[PlanRelationshipUpdate] = []
    confirmed: bool | None = None


@router.post("/runs")
def create_run(body: CreateRunRequest):
    """An empty run, waiting for sources."""
    batch_id = f"run_{uuid.uuid4().hex[:10]}"
    label = body.label or f"Reconciliation {datetime.now(timezone.utc).strftime('%d %b %H:%M')}"
    store.create_batch(batch_id, label, "upload", 0)
    audit.append_event(
        transaction_id=batch_id, event_type="RUN_CREATED", prior_state=None, new_state="DRAFT",
        payload={"batch_id": batch_id, "label": label},
    )
    return {"run_id": batch_id, "label": label, "status": "DRAFT", "sources": []}


def sample_workspace_files(directory: Path | None = None) -> list[Path]:
    """Every export in the sample workspace, in a stable order.

    Listed from the filesystem on each call. The folder is generated and
    regenerated independently of this code, so anything cached here — a
    file list, a count, a set of expected providers — would be a claim
    about data that had already changed.
    """
    directory = directory or DEMO_WORKSPACE_DIR
    if not directory.is_dir():
        return []
    return sorted(
        path for path in directory.iterdir()
        if path.is_file()
        and not path.name.startswith(("_", "."))
        and path.suffix.lower() not in NON_TABULAR_SUFFIXES
    )


def _source_summary(result: dict) -> dict:
    """The per-file answer, in the same fields an upload reports.

    Deliberately the ingest result rather than a re-description of it:
    the classification, the confidence, and whether it has to be
    confirmed are the values `_ingest_file` produced, not values computed
    a second time for this endpoint. The bulky part of `detection`
    (sample rows, per-column guesses, evidence) is left out and stays
    available on GET /runs/{run_id} and the plan.
    """
    detection = result.get("detection") or {}
    return {
        "source_id": result["source_id"],
        "filename": result["filename"],
        "source_type": result["source_type"],
        "role": result["role"],
        "row_count": detection.get("row_count"),
        "detected_source_type": result.get("detected_source_type"),
        "detection_confidence": result.get("detection_confidence"),
        "provider": result.get("provider"),
        "needs_confirmation": bool(result.get("needs_confirmation")),
        "role_confirmed": bool(result.get("role_confirmed")),
        "duplicate_of": result.get("duplicate_of"),
        "duplicate_of_filename": result.get("duplicate_of_filename"),
        "unmapped_required": detection.get("unmapped_required") or [],
        "needs_user_input": bool(result.get("needs_user_input")),
        "format": result.get("format"),
        "truncated": bool(result.get("truncated")),
        "stage": result.get("stage"),
        "stage_label": result.get("stage_label"),
        "suggested_role": result.get("suggested_role"),
    }


@router.post("/runs/sample")
def create_sample_run(body: CreateRunRequest | None = None):
    """Load the sample workspace as a run — one click, nothing to upload.

    Every file goes through the same path an upload takes: read the
    bytes, detect the schema, classify the source, save it. Nothing is
    pre-labelled and nothing is injected past the classifier, so what
    this returns is what the classifier actually made of those files. A
    file it is not confident about comes back `needs_confirmation` and
    blocks execution exactly as an uploaded one would — the sample is a
    shortcut past finding files, not past the parts that ask.
    """
    files = sample_workspace_files()
    if not files:
        raise HTTPException(
            404,
            f"no sample workspace on disk at {DEMO_WORKSPACE_DIR.name}/ — "
            "generate it with data/generate_demo_workspace.py",
        )

    label = (body.label if body else None) or (
        f"Sample workspace {datetime.now(timezone.utc).strftime('%d %b %H:%M')}"
    )
    created = create_run(CreateRunRequest(label=label))
    run_id = created["run_id"]

    results: list[dict] = []
    errors: list[dict] = []
    used = 0

    for path in files:
        if len(results) >= MAX_SOURCES_PER_RUN:
            errors.append({"ok": False, "filename": path.name, "status": 400,
                           "error": f"over the {MAX_SOURCES_PER_RUN}-file limit for one workspace"})
            continue
        try:
            raw = path.read_bytes()
            if used + len(raw) > MAX_WORKSPACE_BYTES:
                raise _FileRejected(
                    413,
                    f"this workspace would exceed its total size limit of "
                    f"{MAX_WORKSPACE_BYTES // (1024 * 1024)}MB",
                )
            # `None` for the declared type is what an upload sends when it
            # wants the file classified rather than asserted.
            results.append(_ingest_file(run_id, path.name, raw, None))
            used += len(raw)
        except _FileRejected as exc:
            errors.append({"ok": False, "filename": path.name, "status": exc.status, "error": exc.detail})
        except OSError as exc:
            errors.append({"ok": False, "filename": path.name, "status": 400,
                           "error": f"could not be read from disk: {type(exc).__name__}: {exc}"})

    sources = [_source_summary(r) for r in results]
    return {
        **created,
        "sources": sources,
        "source_count": len(sources),
        # Rows ingested across every source, both sides. This is what was
        # read out of the files, not what a reconciliation will produce —
        # that number does not exist until the run is executed.
        "record_count": sum(int(s["row_count"] or 0) for s in sources),
        "errors": errors,
        "file_count": store.count_sources(run_id),
        "max_files": MAX_SOURCES_PER_RUN,
        "needs_confirmation": [s["filename"] for s in sources if s["needs_confirmation"]],
        "duplicates": [
            {"filename": s["filename"], "duplicate_of_filename": s["duplicate_of_filename"]}
            for s in sources if s["duplicate_of"]
        ],
        "workspace": DEMO_WORKSPACE_DIR.name,
    }


@router.get("/runs")
def list_runs(limit: int = 20, include_drafts: bool = False):
    # Drafts are hidden by default. A workspace someone created and never
    # executed is not a failed run, and a list dominated by them makes the
    # product look like it keeps breaking — but they are still real, so
    # `include_drafts=true` returns them rather than the data disappearing.
    return {"runs": store.list_batches(limit=limit, include_drafts=include_drafts)}


@router.get("/runs/{run_id}")
def get_run(run_id: str):
    batch = store.get_batch(run_id)
    if not batch:
        raise HTTPException(404, "run not found")
    batch["outcome_counts"] = store.batch_outcome_counts(run_id)
    batch["sources"] = store.list_sources(run_id)
    return batch


class _FileRejected(Exception):
    """One file in a multi-file upload could not be ingested."""

    def __init__(self, status: int, detail: str) -> None:
        super().__init__(detail)
        self.status = status
        self.detail = detail


def _identifier_sample(rows: list[dict], mapping: dict[str, str]) -> list[str]:
    """A bounded sample of this file's cross-system identifiers.

    Used only to propose relationships between files on *observed*
    overlap. It is a sample, so its absence proves nothing and every
    place it is reported says so.
    """
    columns = [mapping.get(c) for c in ("reference", "transaction_id")]
    columns = [c for c in columns if c]
    if not columns:
        return []
    seen: list[str] = []
    unique: set[str] = set()
    for row in rows[:IDENTIFIER_SAMPLE_ROWS]:
        for column in columns:
            value = str(row.get(column, "") or "").strip().upper()
            if value and value not in unique:
                unique.add(value)
                seen.append(value)
                if len(seen) >= IDENTIFIER_SAMPLE_SIZE:
                    return seen
    return seen


def _ingest_file(run_id: str, filename: str, raw: bytes, requested_type: str | None) -> dict:
    """One uploaded file: read it, classify it, store it, report it.

    Raises `_FileRejected` rather than an HTTPException so a bad file in
    a batch of twelve fails on its own line instead of taking the other
    eleven with it. A single-file upload turns the same failure back into
    the HTTP error it always was.
    """
    if len(raw) > MAX_UPLOAD_BYTES:
        raise _FileRejected(413, f"file exceeds {MAX_UPLOAD_BYTES // (1024 * 1024)}MB")

    try:
        read = read_table(filename, raw)
    except UnreadableFile as exc:
        raise _FileRejected(400, str(exc)) from exc

    if not read.columns:
        raise _FileRejected(400, "no columns found — is this a CSV?")
    if not read.rows:
        raise _FileRejected(400, "file has headers but no rows")

    detected = detect_schema(read.columns, read.rows)
    classification = classify_source(filename, read.columns, read.rows, detected)

    # An explicitly declared type is a statement by the person uploading,
    # so it stands and needs no confirmation. AUTO means the classifier
    # decides — and then a low-confidence answer must be confirmed before
    # it can be reconciled on.
    explicit = bool(requested_type) and requested_type.upper() != AUTO
    if explicit:
        kind = SourceType(requested_type)                   # validated by the caller
        role_confirmed = True
    else:
        kind = classification.source_type
        role_confirmed = not classification.needs_confirmation

    content_hash = hashlib.sha256(raw).hexdigest()
    duplicate = store.find_source_by_hash(run_id, content_hash)

    source_id = f"src_{uuid.uuid4().hex[:10]}"
    detection = {
        "columns": detected.columns,
        "row_count": detected.row_count,
        "amount_scale": detected.amount_scale,
        "unmapped_required": detected.unmapped_required,
        "debit_column": detected.debit_column,
        "credit_column": detected.credit_column,
        "sample_rows": detected.sample_rows,
        "guesses": [
            {"column": g.column, "canonical": g.canonical, "confidence": g.confidence,
             "reason": g.reason, "samples": g.samples}
            for g in detected.guesses
        ],
        "classification": classification.to_dict(),
        "identifier_sample": _identifier_sample(read.rows, detected.mapping),
        "role_confirmed": role_confirmed,
        "role_declared_by": "user" if explicit else "detection",
        "format": read.fmt,
        "header_row": read.header_row or 1,
        "sheet_name": read.sheet_name,
        "truncated": read.truncated,
        "read_notes": read.notes,
        "content_hash": content_hash,
        "duplicate_of": duplicate["source_id"] if duplicate else None,
    }

    store.save_source(source_id, run_id, filename, kind.value, kind.role,
                      detected.row_count, detected.mapping, detection, read.csv_text,
                      content_hash=content_hash)
    audit.append_event(
        transaction_id=run_id, event_type="SOURCE_UPLOADED", prior_state="DRAFT", new_state="DRAFT",
        payload={"batch_id": run_id, "source_id": source_id, "filename": filename,
                 "source_type": kind.value, "role": kind.role, "rows": detected.row_count,
                 "unmapped_required": detected.unmapped_required,
                 "detected_source_type": classification.source_type.value,
                 "detection_confidence": round(classification.confidence, 3),
                 "provider": classification.provider,
                 "role_declared_by": detection["role_declared_by"],
                 "duplicate_of": detection["duplicate_of"],
                 "content_hash": content_hash},
    )

    result = {
        "ok": True,
        "source_id": source_id,
        "filename": filename,
        "source_type": kind.value,
        "role": kind.role,
        "canonical_fields": list(CANONICAL_FIELDS),
        "mapping": detected.mapping,
        "detection": detection,
        "needs_user_input": detected.needs_user_input,
        "role_confirmed": role_confirmed,
        "duplicate_of": detection["duplicate_of"],
        "duplicate_of_filename": duplicate["filename"] if duplicate else None,
        "format": read.fmt,
        "truncated": read.truncated,
        "read_notes": read.notes,
    }
    result.update(classification.to_dict())
    # The declared type wins over the detected one in the top-level
    # `source_type`; both are reported so a user can see they differ.
    result["source_type"] = kind.value
    result["role"] = kind.role
    return result


@router.post("/runs/{run_id}/sources")
async def upload_sources(
    run_id: str,
    files: list[UploadFile] = File(default=[]),
    file: list[UploadFile] = File(default=[]),
    source_type: str | None = Form(default=None),
    source_types: list[str] = Form(default=[]),
):
    """Take any number of files, work out what each one is, and report it.

    Both field names are supported: `files` (one or many) and `file`
    (what the single-file callers already send). A list of one is the
    same thing as one, so a single-file upload still gets its result at
    the top level of the response as well as inside `sources`.

    `source_type` is optional. Omitted — or sent as "AUTO" — means the
    file is classified from its contents, and a low-confidence answer
    comes back with `needs_confirmation`, which blocks execution until a
    person confirms it. `source_types` may carry one declared type per
    file, in upload order.
    """
    if not store.get_batch(run_id):
        raise HTTPException(404, "run not found")

    uploads = list(files) + list(file)
    if not uploads:
        raise HTTPException(400, "send at least one file, as 'files' (one or many) or 'file'")

    declared: list[str | None] = list(source_types) if source_types else []
    if len(declared) not in (0, len(uploads)):
        raise HTTPException(400, f"{len(declared)} source_types for {len(uploads)} files — send one per file, or none")

    def _validate(value: str | None) -> str | None:
        if value is None or value.upper() == AUTO:
            return None
        try:
            SourceType(value)
        except ValueError:
            raise HTTPException(400, f"unknown source type '{value}'") from None
        return value

    fallback = _validate(source_type)
    per_file = [_validate(v) for v in declared] if declared else [fallback] * len(uploads)

    existing = store.count_sources(run_id)
    if existing + len(uploads) > MAX_SOURCES_PER_RUN:
        raise HTTPException(
            400,
            f"this workspace holds {existing} file(s) and {len(uploads)} more were sent, over the "
            f"limit of {MAX_SOURCES_PER_RUN}. Remove sources, or split the reconciliation into "
            f"separate runs.",
        )

    used = store.workspace_bytes(run_id)
    results: list[dict] = []
    errors: list[dict] = []

    for index, upload in enumerate(uploads):
        filename = upload.filename or f"upload-{index + 1}.csv"
        try:
            raw = await upload.read()
            if used + len(raw) > MAX_WORKSPACE_BYTES:
                raise _FileRejected(
                    413,
                    f"this workspace would exceed its total size limit of "
                    f"{MAX_WORKSPACE_BYTES // (1024 * 1024)}MB",
                )
            result = _ingest_file(run_id, filename, raw, per_file[index])
            used += len(raw)
            results.append(result)
        except _FileRejected as exc:
            errors.append({"ok": False, "filename": filename, "status": exc.status, "error": exc.detail})

    # A single file that failed is the error it always was; a batch where
    # everything failed is reported the same way rather than as an empty
    # success.
    if errors and not results:
        raise HTTPException(errors[0]["status"], errors[0]["error"])

    response = {
        "run_id": run_id,
        "count": len(results),
        "sources": results,
        "errors": errors,
        "file_count": store.count_sources(run_id),
        "max_files": MAX_SOURCES_PER_RUN,
    }
    if len(uploads) == 1 and results:
        # Backwards compatibility: the single-file shape is preserved for
        # existing callers, which read source_id/mapping/detection at the
        # top level.
        response = {**results[0], **response}
    return response


@router.put("/runs/{run_id}/sources/{source_id}/mapping")
def update_mapping(run_id: str, source_id: str, body: MappingRequest):
    """Correct what detection guessed.

    Re-parsed against the stored file so the preview reflects the user's
    mapping rather than the original guess.
    """
    source = store.get_source(source_id)
    if not source or source["batch_id"] != run_id:
        raise HTTPException(404, "source not found in this run")

    unknown = [f for f in body.mapping if f not in CANONICAL_FIELDS]
    if unknown:
        raise HTTPException(400, f"unknown canonical field(s): {', '.join(unknown)}")

    columns = source["detection"]["columns"]
    bad = [c for c in body.mapping.values() if c and c not in columns]
    if bad:
        raise HTTPException(400, f"column(s) not present in this file: {', '.join(bad)}")

    kind = SourceType(body.source_type) if body.source_type else SourceType(source["source_type"])
    mapping = {k: v for k, v in body.mapping.items() if v}
    store.update_source_mapping(source_id, mapping, kind.value, kind.role)

    detection = source["detection"]
    if body.amount_scale in ("major", "minor"):
        detection["amount_scale"] = body.amount_scale
    detection["unmapped_required"] = [f for f in ("amount", "date") if f not in mapping]
    if body.source_type:
        # Stating the type here is the confirmation the classifier asked
        # for, so the run is no longer blocked on it.
        detection["role_confirmed"] = True
        detection["role_declared_by"] = "user"
    store.save_source(source_id, run_id, source["filename"], kind.value, kind.role,
                      source["row_count"], mapping, detection, source["raw_csv"],
                      content_hash=source.get("content_hash"))
    return {"source_id": source_id, "mapping": mapping, "source_type": kind.value,
            "role": kind.role, "unmapped_required": detection["unmapped_required"],
            "role_confirmed": bool(detection.get("role_confirmed"))}


@router.delete("/runs/{run_id}/sources/{source_id}")
def remove_source(run_id: str, source_id: str):
    source = store.get_source(source_id)
    if not source or source["batch_id"] != run_id:
        raise HTTPException(404, "source not found in this run")
    store.delete_source(source_id)
    return {"source_id": source_id, "removed": True}


@router.get("/runs/{run_id}/plan")
def get_plan(run_id: str):
    """The workspace as a money-flow map, gaps included.

    Read the module docstring in `app/ingest/flow.py` before wiring this
    into a diagram: the stages are a plan and a provenance view over a
    two-sided engine, not evidence of hop-by-hop matching.
    """
    batch = store.get_batch(run_id)
    if not batch:
        raise HTTPException(404, "run not found")
    sources = store.list_sources(run_id)
    records = store.list_record_provenance(run_id)
    return flow.build_plan(batch, sources, saved_plan=store.get_run_plan(run_id), records=records)


@router.put("/runs/{run_id}/plan")
def update_plan(run_id: str, body: PlanRequest):
    """Confirm or correct the plan: what each file is, and which pairs relate.

    Confirming a source's role is the gate the classifier defers to — a
    file it was not sure about cannot be reconciled on until this has
    been answered.
    """
    batch = store.get_batch(run_id)
    if not batch:
        raise HTTPException(404, "run not found")

    by_id = {s["source_id"]: s for s in store.list_sources(run_id, include_raw=True)}
    unknown = [u.source_id for u in body.sources if u.source_id not in by_id]
    unknown += [
        sid for r in body.relationships for sid in (r.from_source_id, r.to_source_id)
        if sid not in by_id
    ]
    if unknown:
        raise HTTPException(404, f"source(s) not in this run: {', '.join(sorted(set(unknown)))}")

    saved = store.get_run_plan(run_id)
    saved_sources = dict(saved.get("sources") or {})
    saved_relationships = dict(saved.get("relationships") or {})

    for update in body.sources:
        source = by_id[update.source_id]
        detection = source["detection"]
        kind = SourceType(source["source_type"])
        if update.source_type:
            try:
                kind = SourceType(update.source_type)
            except ValueError:
                raise HTTPException(400, f"unknown source type '{update.source_type}'") from None
        role = update.role or kind.role
        if role not in ("LEDGER", "SETTLEMENT"):
            raise HTTPException(400, f"role must be LEDGER or SETTLEMENT, got '{role}'")
        detection["role_confirmed"] = bool(update.confirmed)
        detection["role_declared_by"] = "user" if update.confirmed else "detection"
        store.save_source(source["source_id"], run_id, source["filename"], kind.value, role,
                          source["row_count"], source["mapping"], detection, source["raw_csv"],
                          content_hash=source.get("content_hash"))
        saved_sources[update.source_id] = {
            "source_type": kind.value, "role": role, "confirmed": bool(update.confirmed),
            "note": update.note,
        }

    for relationship in body.relationships:
        key = f"{relationship.from_source_id}->{relationship.to_source_id}"
        saved_relationships[key] = {"confirmed": relationship.confirmed, "note": relationship.note}

    plan_state = {
        "sources": saved_sources,
        "relationships": saved_relationships,
        "confirmed": bool(body.confirmed) if body.confirmed is not None else bool(saved.get("confirmed")),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    store.save_run_plan(run_id, plan_state)
    audit.append_event(
        transaction_id=run_id, event_type="PLAN_UPDATED", prior_state="DRAFT", new_state="DRAFT",
        payload={"batch_id": run_id,
                 "sources_confirmed": [u.source_id for u in body.sources if u.confirmed],
                 "relationships": list(saved_relationships.keys()),
                 "confirmed": plan_state["confirmed"]},
    )

    sources = store.list_sources(run_id)
    records = store.list_record_provenance(run_id)
    return flow.build_plan(store.get_batch(run_id), sources, saved_plan=plan_state, records=records)


# ---------------------------------------------------------------------------
# Execution progress
# ---------------------------------------------------------------------------

# The phases the executor writes as it crosses them. `DECIDING` is one
# phase on purpose: deterministic matching and AI investigation are
# interleaved inside a single pass over the records — the model is
# consulted mid-record, when that record's evidence turns out to be
# ambiguous — so presenting them as two sequential passes would be a
# picture of a pipeline this system does not have.
PHASE_INSPECTING = "INSPECTING"
PHASE_NORMALISING = "NORMALISING"
PHASE_DECIDING = "DECIDING"
PHASE_REVIEW = "REVIEW"
PHASE_COMPLETE = "COMPLETE"

# key, label, and the fixed part of the description.
PIPELINE_STAGES: tuple[tuple[str, str, str], ...] = (
    ("INSPECTING", "Inspect sources",
     "Each file is read, its columns detected and its type classified."),
    ("NORMALISING", "Normalise records",
     "Rows from every source are folded into canonical ledger and settlement records."),
    ("DETERMINISTIC", "Deterministic matching",
     "Every record is decided on exact and corroborated evidence first."),
    ("INVESTIGATING", "AI investigation",
     "Consulted only where the evidence is ambiguous. Runs inside the matching pass, "
     "record by record — not as a separate stage afterwards."),
    ("REVIEW", "Human review",
     "Whatever could not be resolved is held for a person."),
)

_PENDING, _ACTIVE, _DONE = "PENDING", "ACTIVE", "DONE"

# Which stage rows are PENDING / ACTIVE / DONE in each real phase. Three
# rows are ACTIVE together in DECIDING because three things are genuinely
# advancing together there: records are being decided, some of them are
# being investigated as they are decided, and the ones that resolve to
# neither are landing in review as it happens.
_STAGE_STATES: dict[str, tuple[str, ...]] = {
    "": (_PENDING, _PENDING, _PENDING, _PENDING, _PENDING),
    PHASE_INSPECTING: (_ACTIVE, _PENDING, _PENDING, _PENDING, _PENDING),
    PHASE_NORMALISING: (_DONE, _ACTIVE, _PENDING, _PENDING, _PENDING),
    PHASE_DECIDING: (_DONE, _DONE, _ACTIVE, _ACTIVE, _ACTIVE),
    PHASE_REVIEW: (_DONE, _DONE, _DONE, _DONE, _ACTIVE),
    PHASE_COMPLETE: (_DONE, _DONE, _DONE, _DONE, _DONE),
}


def _derive_phase(batch: dict, recorded: dict) -> str:
    """Where the run is, preferring what the executor recorded.

    The recorded phase is the truthful one — it was written by the code
    that was doing the work. It can be absent for two honest reasons: the
    run has never been executed, or it is a dataset-driven batch from
    `/batch/run`, which does not go through this executor at all. Both
    degrade to what the batch row itself can prove.
    """
    phase = str(recorded.get("phase") or "")
    if phase in _STAGE_STATES and phase:
        # A run whose thread died mid-flight would sit in DECIDING for
        # ever; the batch row is the authority on completion.
        if batch.get("status") == "COMPLETED":
            return PHASE_COMPLETE
        return phase

    if batch.get("status") == "COMPLETED":
        return PHASE_COMPLETE
    total = int(batch.get("total_records") or 0)
    processed = int(batch.get("processed_records") or 0)
    if total <= 0 and processed <= 0:
        return ""                                   # nothing has run
    if total > 0 and processed >= total:
        return PHASE_REVIEW
    return PHASE_DECIDING


def _investigating_detail(tier: int, model: int, heuristic: int, unanswered: int) -> str:
    """What actually served the ambiguous records, said plainly.

    The stage is called "AI investigation" and its count is the number of
    records a model answered. When the offline verifier served them
    instead, that count is zero and this sentence has to be the thing
    that stops the row reading as "nothing happened" — and, more
    importantly, stops the run reading as model-assisted when it was not.
    """
    if not tier:
        return "No record needed more than deterministic evidence."

    parts: list[str] = []
    if model:
        parts.append(f"{model} record(s) were ambiguous enough to consult the model.")
    if heuristic:
        parts.append(
            f"{heuristic} record(s) reached the ambiguity tier and were resolved by the offline "
            "classifier — no model was called."
            if not model else
            f"{heuristic} more were resolved by the offline classifier, with no model call."
        )
    if unanswered:
        parts.append(
            f"{unanswered} reached a provider that did not answer and went to human review."
        )
    return " ".join(parts)


def _progress_payload(batch: dict) -> dict:
    """The pipeline's real position, assembled from stored state only.

    Every number here is either a column the executor wrote or a count
    taken from the rows themselves at the moment of the request. Where a
    number is genuinely not yet knowable — how many records a run will
    have before its sources are mapped, how many will need investigating
    before they are decided — the answer is `null`, because a plausible
    stand-in on a progress display is indistinguishable from a real
    measurement and would be trusted like one.
    """
    run_id = batch["batch_id"]
    recorded = store.get_run_progress(run_id)
    counts = store.run_progress_counts(run_id)
    phase = _derive_phase(batch, recorded)
    states = _STAGE_STATES[phase]
    started = bool(phase)

    processed = int(batch.get("processed_records") or 0)
    sized = int(batch.get("total_records") or 0) > 0
    total = int(batch["total_records"]) if sized else None

    # Reaching the semantic tier is not the same thing as calling a
    # model. When no provider is configured the offline heuristic serves
    # that tier, and reporting those as "AI consulted" would tell a
    # viewer a model was used when none was. So `ai_consulted` counts
    # only records a real provider answered, and the tier total is
    # reported separately under its own name.
    ai_consulted = counts["ai_consulted"]
    heuristic_consulted = counts["heuristic_consulted"]
    semantic_tier = counts["semantic_tier_invoked"]
    # Reached the tier, and nothing answered — every provider in the
    # chain failed and the record went to human review.
    unanswered = max(semantic_tier - ai_consulted - heuristic_consulted, 0)
    if not semantic_tier:
        semantic_backend = None
    elif ai_consulted and heuristic_consulted:
        semantic_backend = "mixed"
    elif ai_consulted:
        semantic_backend = "model"
    elif heuristic_consulted:
        semantic_backend = "heuristic"
    else:
        semantic_backend = None                 # reached, answered by nothing
    unresolved = counts["unresolved"]

    sources_total = recorded.get("sources_total")
    sources_done = recorded.get("sources_done")
    ledger = recorded.get("ledger_records")
    settlements = recorded.get("settlement_records")
    canonical = None if ledger is None or settlements is None else ledger + settlements

    details = {
        "INSPECTING": (
            f"{sources_done} of {sources_total} sources read, mapped and classified."
            if sources_done is not None and sources_total else None
        ),
        "NORMALISING": (
            f"{ledger} ledger and {settlements} settlement records built from those files."
            if canonical is not None else None
        ),
        "DETERMINISTIC": (
            f"{processed} of {total} records decided." if total is not None else None
        ),
        "INVESTIGATING": _investigating_detail(
            semantic_tier, ai_consulted, heuristic_consulted, unanswered
        ) if states[3] != _PENDING else None,
        "REVIEW": (
            f"{unresolved} record(s) could not be resolved automatically."
            if states[4] != _PENDING else None
        ),
    }
    # A stage that has not started has not measured anything, so it
    # reports nothing rather than a zero that reads like a result.
    values: dict[str, tuple[int | None, int | None]] = {
        "INSPECTING": (sources_done, sources_total),
        "NORMALISING": (canonical, None),
        "DETERMINISTIC": (processed, total),
        "INVESTIGATING": (ai_consulted, None),
        "REVIEW": (unresolved, None),
    }

    stages = []
    for (key, label, description), state in zip(PIPELINE_STAGES, states):
        count, stage_total = values[key] if state != _PENDING else (None, None)
        detail = details[key] if state != _PENDING else None
        stages.append({
            "key": key,
            "label": label,
            "detail": detail or description,
            "state": state,
            "count": count,
            "total": stage_total,
        })

    # The headline. DECIDING is reported as INVESTIGATING once a model
    # has actually been consulted on this run and as DETERMINISTIC before
    # that — both describe the same single pass, named for what it is
    # currently doing rather than promoted to a phase of its own. A run
    # the offline verifier is serving never reports INVESTIGATING, because
    # nothing is being investigated by a model.
    if phase == PHASE_DECIDING:
        stage = "INVESTIGATING" if ai_consulted > 0 else "DETERMINISTIC"
    elif phase == "":
        stage = "INSPECTING"                        # the stage that will run first
    else:
        stage = phase

    return {
        "run_id": run_id,
        "stage": stage,
        "stages": stages,
        "processed": processed,
        "total": total,
        # Records a real provider answered. Never includes work the
        # offline verifier did — see `semantic_tier_invoked` for the
        # total that reached the ambiguity tier at all, and
        # `semantic_backend` for what served it.
        "ai_consulted": ai_consulted,
        "semantic_tier_invoked": semantic_tier,
        "heuristic_consulted": heuristic_consulted,
        "semantic_backend": semantic_backend,
        "unresolved": unresolved,
        # `batches.status` is deliberately not echoed here: a run is
        # created with status RUNNING before it has any sources, so it
        # would say "running" about a workspace nobody has executed.
        # These two are the ones that are true.
        "started": started,
        "complete": phase == PHASE_COMPLETE,
        "updated_at": recorded.get("updated_at"),
        # Set only if the execution thread stopped on an error. Null is
        # the normal case; a run that failed says so instead of sitting
        # at a stage for ever.
        "error": recorded.get("error"),
    }


@router.get("/runs/{run_id}/progress")
def get_run_progress(run_id: str):
    """Where this run actually is.

    Cheap enough to poll: one row from `batches` and one aggregate over
    that batch's records.
    """
    batch = store.get_batch(run_id)
    if not batch:
        raise HTTPException(404, "run not found")
    return _progress_payload(batch)


@router.post("/runs/{run_id}/execute")
def execute_run(run_id: str, body: ExecuteRequest):
    """Map every source, then reconcile.

    Refuses rather than guesses when a run cannot produce two sides: a
    reconciliation with nothing to reconcile against would return a page
    of exceptions that say more about the upload than about the books.
    """
    batch = store.get_batch(run_id)
    if not batch:
        raise HTTPException(404, "run not found")
    sources = store.list_sources(run_id, include_raw=True)
    if not sources:
        raise HTTPException(400, "upload at least one source before running")

    blocked = [s["filename"] for s in sources if s["detection"].get("unmapped_required")]
    if blocked:
        raise HTTPException(400, f"map the required columns first for: {', '.join(blocked)}")

    # A role that was inferred rather than stated, and inferred weakly,
    # is not a basis for reconciling money. Putting an order book on the
    # settlement side would compare the ledger against itself and return
    # a page of clean matches, so the guess has to be confirmed first.
    unconfirmed = [
        s["filename"] for s in sources
        if (s["detection"].get("classification") or {}).get("needs_confirmation")
        and not s["detection"].get("role_confirmed")
    ]
    if unconfirmed:
        raise HTTPException(
            400,
            "confirm what these file(s) are before running — detection was not confident enough: "
            + ", ".join(unconfirmed),
        )

    # From here on the executor records where it is, at each boundary it
    # genuinely crosses. Nothing below writes a phase it is not in.
    store.save_run_progress(run_id, {
        "phase": PHASE_INSPECTING,
        "sources_total": len(sources),
        "sources_done": 0,
        "started_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    })

    mapped = []
    for source in sources:
        _, rows = parse_csv(source["raw_csv"])
        detection = source["detection"]
        result = map_rows(
            rows, source["mapping"], SourceType(source["source_type"]), source["source_id"],
            detection.get("amount_scale", "major"),
            detection.get("debit_column"), detection.get("credit_column"),
            filename=source["filename"],
            header_offset=int(detection.get("header_row") or 1),
        )
        store.record_source_outcome(source["source_id"], result.accepted_count, len(result.rejected))
        mapped.append(result)
        store.update_run_progress(run_id, sources_done=len(mapped))

    ledger, settlements, rejected = combine(mapped)
    ledger_provenance, settlement_provenance = combine_provenance(mapped)
    store.update_run_progress(
        run_id, phase=PHASE_NORMALISING, ledger_records=len(ledger),
        settlement_records=len(settlements), rejected_rows=len(rejected),
    )
    # A run that cannot produce two sides stops here. The refusal is
    # recorded on the run so a progress view shows a run that stopped and
    # why, rather than one that appears to still be working.
    if not ledger:
        detail = "no ledger-side records — add an orders or accounting source"
        store.update_run_progress(run_id, error=detail)
        raise HTTPException(400, detail)
    if not settlements:
        detail = "no settlement-side records — add a gateway or bank source"
        store.update_run_progress(run_id, error=detail)
        raise HTTPException(400, detail)

    records = [ReconciliationRecord(record_id=r.order_id, merchant=r) for r in ledger]

    # Provenance is index-aligned with `ledger` (see combine_provenance),
    # so a record's origin is its position, not a lookup by id — two
    # sources can legitimately carry the same order id, and keying by it
    # would attribute one file's row to the other file.
    settlement_origin = {
        record.payment_id: origin
        for record, origin in zip(settlements, settlement_provenance)
    }
    label = body.label or batch["label"]
    store.set_batch_total(run_id, len(records), label)
    audit.append_event(
        transaction_id=run_id, event_type="BATCH_STARTED", prior_state="DRAFT", new_state="RUNNING",
        payload={"batch_id": run_id, "dataset": "upload", "total": len(records),
                 "ledger_records": len(ledger), "settlement_records": len(settlements),
                 "rejected_rows": len(rejected), "sources": len(sources)},
    )

    def _run():
        index = ReferenceIndex(settlements)
        policy = PolicyConfig()
        verifier = get_semantic_verifier()

        def _provenance(i: int, result) -> dict:
            origin = {"ledger": ledger_provenance[i] if i < len(ledger_provenance) else None}
            if result.matched_payment_id:
                origin["settlement"] = settlement_origin.get(result.matched_payment_id)
            return origin

        def on_record(i, total, record, result):
            store.save_record(run_id, i, record, result, index.exact_candidates(record.merchant),
                              provenance=_provenance(i, result))
            store.increment_batch_progress(run_id)
            audit.append_event(
                transaction_id=record.record_id, event_type="RECORD_DECIDED",
                prior_state=None, new_state=result.outcome.value,
                payload={"batch_id": run_id, "seq": i, "total": total, "record_id": record.record_id,
                         "outcome": result.outcome.value, "reason": result.reason,
                         "ai_invoked": result.ai_invoked},
            )
            # The last per-record callback is the end of the matching
            # pass. What follows inside process_batch is the cross-record
            # integrity work — duplicate claims and aggregated
            # settlements — which only routes records to human review, so
            # the run is genuinely in its review phase from here.
            if i + 1 >= total:
                store.update_run_progress(run_id, phase=PHASE_REVIEW)

        def on_revision(i, record, result):
            store.save_record(run_id, i, record, result, index.exact_candidates(record.merchant),
                              provenance=_provenance(i, result))
            audit.append_event(
                transaction_id=record.record_id, event_type="RECORD_REVISED",
                prior_state="RECONCILED", new_state=result.outcome.value,
                payload={"batch_id": run_id, "record_id": record.record_id,
                         "outcome": result.outcome.value, "reason": result.reason},
            )

        try:
            process_batch(records, settlements, policy=policy, semantic_verifier=verifier,
                          on_record=on_record, on_revision=on_revision)
        except Exception as exc:  # noqa: BLE001 — a thread that dies silently leaves the run "in progress" for ever
            store.update_run_progress(run_id, error=f"{type(exc).__name__}: {exc}")
            audit.append_event(
                transaction_id=run_id, event_type="BATCH_FAILED", prior_state="RUNNING",
                new_state="RUNNING",
                payload={"batch_id": run_id, "error": f"{type(exc).__name__}: {exc}"},
            )
            raise
        store.mark_batch_complete(run_id)
        store.update_run_progress(run_id, phase=PHASE_COMPLETE)
        audit.append_event(
            transaction_id=run_id, event_type="BATCH_COMPLETED", prior_state="RUNNING",
            new_state="COMPLETED", payload={"batch_id": run_id, "total": len(records)},
        )

    store.update_run_progress(run_id, phase=PHASE_DECIDING)
    threading.Thread(target=_run, daemon=True).start()
    return {
        "run_id": run_id, "total_records": len(records), "label": label,
        "ledger_records": len(ledger), "settlement_records": len(settlements),
        "rejected_rows": rejected[:50], "rejected_count": len(rejected),
    }


@router.get("/runs/{run_id}/export")
def export_run(run_id: str, outcome: str | None = None, format: str = "csv"):
    """Results as a spreadsheet, including the evidence behind each decision.

    Two formats, because reconciliation output lands in a spreadsheet and
    half of finance opens CSV in Excel while the other half wants a
    workbook that already has its header frozen. Both carry exactly the
    same columns — the evidence (reason, explanation, rejected candidates,
    source provenance) is what makes the file worth having, so it is never
    trimmed for the "convenient" format.

    `outcome` narrows the export to one outcome so the file matches the
    filter the operator is looking at on screen.
    """
    fmt = (format or "csv").strip().lower()
    if fmt not in ("csv", "xlsx"):
        raise HTTPException(400, f"unknown export format '{format}' — use 'csv' or 'xlsx'")
    if not store.get_batch(run_id):
        raise HTTPException(404, "run not found")
    if outcome:
        outcome = outcome.strip().upper()

    rows = store.list_records(run_id, outcome=outcome, limit=100_000)
    header, body = _run_export_table(rows)
    scope = EXPORT_SCOPE.get(outcome or "", "all-records")
    return table_response(header, body, filename=f"{run_id}-{scope}", sheet_title="Reconciliation",
                          fmt=fmt)


# What each outcome filter is called in the downloaded file's name, so a
# folder of exports is readable without opening them.
EXPORT_SCOPE = {
    "": "all-records",
    "RECONCILED": "reconciled",
    "EXCEPTION": "exceptions",
    "HUMAN_REVIEW": "human-review",
}


def _run_export_table(rows: list[dict]) -> tuple[list[str], list[list]]:
    """The exported columns, in one place, for every format.

    The reasoning columns are the point of the export: an operator working
    offline has to be able to see *why* a record was decided the way it
    was and *which row of which file* the evidence came from, or the file
    is a list of verdicts with nothing behind them.
    """
    header = [
        "record_id", "outcome", "exception_type", "severity", "matched_payment_id",
        "amount_minor", "currency", "reference", "description", "reason", "explanation",
        "recommended_action", "ai_invoked", "review_state", "rejected_candidates",
        # Appended, not inserted: existing consumers read by position.
        "ledger_source_file", "ledger_source_row",
        "settlement_source_file", "settlement_source_row",
    ]
    body: list[list] = []
    for row in rows:
        merchant = json.loads(row["merchant_json"])
        considered = json.loads(row.get("considered_json") or "[]")
        rejected = "; ".join(
            f"{c['payment_id']} ({c['admissibility_reason']})"
            for c in considered if not c.get("admissible")
        )
        provenance = json.loads(row.get("provenance_json") or "{}") or {}
        ledger_origin = provenance.get("ledger") or {}
        settlement_origin = provenance.get("settlement") or {}
        body.append([
            row["record_id"], row["outcome"], row.get("exception_type") or "",
            row.get("severity") or "", row.get("matched_payment_id") or "",
            merchant.get("amount_minor"), merchant.get("currency"),
            merchant.get("reference_id") or "", merchant.get("description") or "",
            row.get("reason") or "",
            row.get("explanation") or "", row.get("recommended_action") or "",
            "yes" if row.get("ai_invoked") else "no", row.get("review_state") or "OPEN",
            rejected,
            ledger_origin.get("filename") or "", ledger_origin.get("file_row") or "",
            settlement_origin.get("filename") or "", settlement_origin.get("file_row") or "",
        ])
    return header, body


# Column widths are computed from the content, then clamped. Unclamped,
# one long explanation makes a column wider than the screen and the sheet
# is unreadable at the exact moment it matters; too narrow and every
# evidence column shows "###".
_XLSX_MIN_WIDTH = 10
_XLSX_MAX_WIDTH = 52
_XLSX_WIDTH_SAMPLE = 400


def table_response(header: list[str], rows: list[list], *, filename: str, sheet_title: str,
                   fmt: str = "csv"):
    """One table, downloaded as either CSV or a real XLSX workbook.

    Shared by the run export and the review-queue export so the two can
    never drift into carrying different evidence for the same decision.
    """
    import csv
    import io

    from fastapi.responses import StreamingResponse

    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_title[:31]                  # Excel's hard limit
        sheet.append(header)
        for row in rows:
            sheet.append(row)

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center")
        # The header stays put while an operator scrolls 3,000 records —
        # without it the columns are unidentifiable past the first screen.
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for index, name in enumerate(header, start=1):
            widest = len(str(name))
            for row in rows[:_XLSX_WIDTH_SAMPLE]:
                value = row[index - 1] if index - 1 < len(row) else ""
                widest = max(widest, len(str(value if value is not None else "")))
            sheet.column_dimensions[get_column_letter(index)].width = min(
                max(widest + 2, _XLSX_MIN_WIDTH), _XLSX_MAX_WIDTH
            )

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )

    text = io.StringIO()
    writer = csv.writer(text)
    writer.writerow(header)
    for row in rows:
        writer.writerow(row)
    return StreamingResponse(
        iter([text.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
    )
